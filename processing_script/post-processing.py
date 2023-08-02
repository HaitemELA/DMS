import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import pandas as pd

#CIS
cis_path = r'C:\Users\HaitemElAaouani\Documents\DMS\CIS.xlsx'
cis = pd.read_excel(cis_path)

# Path of Excel file
path = r'C:\Users\HaitemElAaouani\Documents\DMS\RawProfiles'
xlsx_file = os.path.join(path, 'GPC_profiles.xlsx')

#Profile to process
uneditedSheet = 'Condition (Unedited)'

# First row formatting Blue	 
blue_fill = PatternFill(start_color='FF6D9EEB', #4A86E8
                           end_color='FF6D9EEB',
                           fill_type='solid')
bold = Font(bold=True)

# First row renaming
RenameColumns = {'element_id': 'Element Name',
                'element_definition_value': 'Description',
                'element_cardinality': 'Cardinality',
                'element_type_code_value': 'Data Type',
                'element_mustSupport_value': 'Must Support'
                }

# WorkBook definition
wkbk = openpyxl.load_workbook(xlsx_file)

# Serch string in row index
def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column
    return cell.column

# Function to delete reds
def delete_any_red(workbook, unedited_sheet, elemIdCol=3):
    edited_sheet = workbook.copy_worksheet(workbook[unedited_sheet])
    edited_sheet.title = edited_sheet.title.replace('(Unedited) Copy', '(Edited)')

    for row_idx in range(edited_sheet.max_row + 1, 0, -1):
        if edited_sheet.cell(row=row_idx, column=elemIdCol).fill.start_color.index == "FFFF0000":
            edited_sheet.delete_rows(row_idx)

    for col_idx in range(edited_sheet.max_column + 1, 0, -1):
        if edited_sheet.cell(row=1, column=col_idx).fill.start_color.index == "FFFF0000":
            edited_sheet.delete_cols(col_idx)
        else:
            edited_sheet.cell(row=1, column=col_idx).fill = blue_fill

        # First row renaming
        if str(edited_sheet.cell(row=1, column=col_idx).value) in RenameColumns:
            edited_sheet.cell(row=1, column=col_idx).value = RenameColumns[str(edited_sheet.cell(row=1, column=col_idx).value)]

    return edited_sheet.title

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def ElementName_FhirTarget(workbook, Edited_sheet, source_col, destination_col):
    sheet = workbook[Edited_sheet]
    #Loops through selected Rows
    for i in range(2,sheet.max_row + 1):
        s = str(sheet.cell(row = i, column = source_col).value)
        sheet.cell(row = i, column = destination_col).value = str(sheet.cell(row = i, column = source_col).value)

        # Value Sets from CIS
        if cis[cis['FHIR target'] == s]['Value sets'].tolist():
            sheet.cell(row = i, column = 9).value = cis[cis['FHIR target'] == s]['Value sets'].tolist()[0]

        # Element Name
        s = s.replace('.',' ')
        s = s.replace('extension:','')
        sheet.cell(row = i, column = source_col).value = s
# Hide Unedited worksheets
def hide_unideted(file_path, workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name.endswith('(Unedited)'):
            worksheet = workbook[sheet_name]
            worksheet.sheet_state = "hidden"

    workbook.save(xlsx_file)

# Find 'element_id' column
elem_id_col = search_value_in_row_index(wkbk[uneditedSheet], 'element_id', row=1)

# Cleansing of the profile
EditSht = delete_any_red(wkbk, uneditedSheet, elem_id_col)
ElementName_FhirTarget(wkbk, EditSht,2,10)


#Rename columns
wkbk[EditSht].insert_cols(2, amount=1)
wkbk[EditSht].cell(row=1, column=1).value = 'Old ID'
wkbk[EditSht].cell(row=1, column=2).value = 'New ID'
wkbk[EditSht].cell(row=1, column=10).value = 'Value Sets'
wkbk[EditSht].cell(row=1, column=11).value = 'FHIR Target STU3'

# Formatting columns
for i in range(1, wkbk[EditSht].max_column + 1):
    wkbk[EditSht].cell(row=1, column=i).fill = blue_fill
    wkbk[EditSht].cell(row=1, column=i).font = bold

# Hide unedited sheets
hide_unideted(xlsx_file, wkbk)
