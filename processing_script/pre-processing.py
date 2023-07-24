from pyexcel.cookbook import merge_all_to_a_book
import pyexcel
import glob
import os
import shutil

import openpyxl
from openpyxl.styles import PatternFill

# Path of CSV files
path = r'C:\Users\HaitemElAaouani\Downloads\RawProfiles'

# Define the columns to keep / rows to delete
White_Columns = ['element_id','element_short_value','element_definition_value','element_cardinality','element_mustSupport_value']
Red_Rows = ['id', 'extension','modifierExtension']

# Temporary directory for temp CSVs
tmp = os.path.join(path, 'tmp')
os.mkdir(tmp)

#Duplicate the CSVs and rename shorter + save them in the temporary directory:
for file in glob.glob(path + '\*.csv'):
    renamed_file = file.split('.')[0].split('-')[-2]
    pyexcel.save_as(file_name=file, dest_file_name= os.path.join(tmp, renamed_file + ".csv"))

#Combine the CSVs in one Excel file:
merge_all_to_a_book(glob.glob(tmp + "\*.csv"), path + "\GPC_raw_profiles.xlsx")

# Delete the temporary directory:
shutil.rmtree(tmp)

# Function to fill columns and rows with red
def red_function(file_path, output_file):
    workbook = openpyxl.load_workbook(file_path)
    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if sheet.cell(row=1, column=col_idx).value not in White_Columns:
                    sheet.cell(row=row_idx, column=col_idx).fill = red_fill
                if sheet.cell(row=row_idx, column=3).value.split('.')[-1] in Red_Rows:
                    sheet.cell(row=row_idx, column=col_idx).fill = red_fill
        sheet.title = sheet.title.split('.')[0] + ' (Unedited)'

    workbook.save(output_file)

# Make Unideted profiles
file_path = path + "\GPC_raw_profiles.xlsx"
Output_file = path + "\GPC_profiles.xlsx"
red_function(file_path, Output_file)